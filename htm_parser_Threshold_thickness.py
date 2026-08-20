import os
import glob
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def extract_threshold_values(html_file_path):
    """
    Извлекает значения толщины порога из HTML-файла CDMAM анализа
    """
    try:
        with open(html_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()

        soup = BeautifulSoup(content, 'html.parser')

        tables = soup.find_all('table')

        for table in tables:
            text = table.get_text()
            if 'Threshold thickness at 62.50 per cent' in text or 'Threshold thickness at 62.50 per cent.' in text:
                rows = table.find_all('tr')

                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) > 0:
                        first_cell_text = cells[0].get_text().strip()
                        if 'Thickness' in first_cell_text and 'Diameter' not in first_cell_text:
                            thickness_values = []
                            for cell in cells[1:]:
                                cell_text = cell.get_text().strip()
                                try:
                                    float(cell_text)
                                    thickness_values.append(cell_text)
                                except ValueError:
                                    continue

                            if thickness_values:
                                return thickness_values[:16]  # Только первые 16 значений

        return None

    except Exception as e:
        return None


def extract_metadata(html_file_path):
    """
    Извлекает метаданные из пути к файлу
    """
    # Парсим путь
    path_parts = html_file_path.split(os.sep)

    # Ищем папку с проекциями
    projections_folder = None
    for part in path_parts:
        if '_grad_' in part and '_projections' in part:
            projections_folder = part
            break

    # Ищем метод
    method = None
    for part in path_parts:
        if part.upper() in ['MINRES', 'MLEM', 'SIRT']:
            method = part.upper()
            break

    # Ищем количество итераций
    iterations = None
    for part in path_parts:
        match = re.search(r'[\\/](\d+)[\\/]', html_file_path.replace(os.sep, '/'))
        if match:
            # Проверяем, что это число и оно не в названии проекций
            num = match.group(1)
            if num not in ['15', '24', '36', '50']:  # Исключаем числа из названий папок
                iterations = int(num)
                break

    # Альтернативный способ поиска итераций
    if iterations is None:
        # Ищем число в конце пути перед .htm
        match = re.search(r'[\\/]([0-9]+)\.[Hh][Tt][Mm]', html_file_path)
        if match:
            iterations = int(match.group(1))

    return {
        'projections': projections_folder,
        'method': method,
        'iterations': iterations
    }


def find_row_in_sheet(sheet, projections_folder, method, iterations):
    """
    Находит строку в листе Excel по проекциям, методу и итерациям
    """
    for row in range(4, sheet.max_row + 1):  # Начинаем с 4 строки
        cell_projections = sheet.cell(row=row, column=4).value  # Колонка D - Угол
        cell_method = sheet.cell(row=row, column=5).value  # Колонка E - Метод
        cell_iterations = sheet.cell(row=row, column=6).value  # Колонка F - Количество итераций

        if (cell_projections == projections_folder and
                cell_method == method and
                cell_iterations == iterations):
            return row

    return None


def process_and_fill_excel(base_directory, template_path, output_path):
    """
    Обрабатывает HTML файлы и заполняет Excel шаблон
    """
    # Загружаем шаблон Excel
    wb = load_workbook(template_path)
    sheet = wb.active

    # Ищем все .htm файлы
    html_files = glob.glob(os.path.join(base_directory, '**', '*.htm'), recursive=True)
    html_files.extend(glob.glob(os.path.join(base_directory, '**', '*.html'), recursive=True))

    # Диаметры
    diameters = ['0.060', '0.080', '0.100', '0.130', '0.160', '0.200', '0.250', '0.310', '0.400', '0.500', '0.630',
                 '0.800', '1.000', '1.250', '1.600', '2.000']

    processed_count = 0
    not_found_count = 0

    for html_file in sorted(html_files):
        # Извлекаем значения толщины
        threshold_values = extract_threshold_values(html_file)

        if threshold_values:
            # Извлекаем метаданные из пути
            metadata = extract_metadata(html_file)

            if metadata['projections'] and metadata['method'] and metadata['iterations']:
                # Находим строку в Excel
                row_num = find_row_in_sheet(sheet, metadata['projections'], metadata['method'], metadata['iterations'])

                if row_num:
                    # Записываем значения толщины в ячейки (начиная с колонки G)
                    for i, value in enumerate(threshold_values):
                        sheet.cell(row=row_num, column=7 + i, value=float(value))

                    processed_count += 1
                    print(
                        f"✓ Обработан: {metadata['projections']} / {metadata['method']} / {metadata['iterations']} итераций")
                else:
                    not_found_count += 1
                    print(f"✗ Не найдена строка в шаблоне: {html_file}")
            else:
                print(f"✗ Не удалось извлечь метаданные: {html_file}")
        else:
            print(f"✗ Не найдены значения толщины: {html_file}")

    # Сохраняем результат
    wb.save(output_path)
    print("\n" + "=" * 60)
    print(f"Обработано файлов: {processed_count}")
    print(f"Не найдено в шаблоне: {not_found_count}")
    print(f"Результат сохранен в: {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    base_dir = r'D:\Маммо_DBT\DBT_processing\no_grid_denoise_vraster'
    template_path = r'D:\Маммо_DBT\шаблон.xlsx'  # Путь к вашему шаблону
    output_path = base_dir + r'\Шаблон_с_таблицами.xlsx'  # Путь для сохранения результата

    if not os.path.exists(base_dir):
        print(f"Ошибка: Директория не найдена: {base_dir}")
    elif not os.path.exists(template_path):
        print(f"Ошибка: Шаблон не найден: {template_path}")
    else:
        process_and_fill_excel(base_dir, template_path, output_path)