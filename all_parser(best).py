import os
import glob
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def extract_iqfinv(html_file_path):
    """Извлекает IQFinv из HTML файла"""
    try:
        with open(html_file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        # Основной паттерн
        pattern = r'<td align="left">\s*IQFinv:\s*</td>\s*<td align="left">\s*([\d.]+)\s*</td>'
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            return float(match.group(1).strip())

        # Альтернативный паттерн
        alt_pattern = r'IQFinv\s*:\s*([\d.]+)'
        alt_match = re.search(alt_pattern, content, re.IGNORECASE)
        if alt_match:
            return float(alt_match.group(1).strip())

        return None
    except:
        return None


def extract_threshold_values(html_file_path):
    """
    Извлекает значения толщины порога из HTML-файла CDMAM анализа
    """
    try:
        with open(html_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            content = file.read()

        soup = BeautifulSoup(content, 'html.parser')
        tables = soup.find_all('table')

        for table in tables:
            text = table.get_text()
            if 'Threshold thickness at 62.50 per cent' in text or 'Threshold thickness at 62.50 per cent.' in text:
                rows = table.find_all('tr')

                for row in rows:
                    cells = row.find_all('td')
                    if len(cells) > 0:
                        first_cell_text = cells[0].get_text().strip()
                        if 'Thickness' in first_cell_text and 'Diameter' not in first_cell_text:
                            thickness_values = []
                            for cell in cells[1:]:
                                cell_text = cell.get_text().strip()
                                try:
                                    float(cell_text)
                                    thickness_values.append(cell_text)
                                except ValueError:
                                    continue

                            if thickness_values:
                                return thickness_values[:16]  # Только первые 16 значений

        return None

    except Exception as e:
        return None


def extract_metadata(html_file_path):
    """
    Извлекает метаданные из пути к файлу
    """
    path_parts = html_file_path.split(os.sep)

    # Ищем папку с проекциями
    projections_folder = None
    projections_index = -1
    for i, part in enumerate(path_parts):
        if '_grad_' in part and '_projections' in part:
            projections_folder = part
            projections_index = i
            break

    # Ищем метод (следующая папка после проекций)
    method = None
    method_index = -1
    if projections_index != -1 and projections_index + 1 < len(path_parts):
        potential_method = path_parts[projections_index + 1]
        if potential_method.upper() in ['MINRES', 'MLEM', 'SIRT']:
            method = potential_method.upper()
            method_index = projections_index + 1

    # Ищем количество итераций (следующая папка после метода)
    iterations = None
    if method_index != -1 and method_index + 1 < len(path_parts):
        potential_iterations = path_parts[method_index + 1]
        if potential_iterations.isdigit():
            iterations = int(potential_iterations)

    # Если не нашли итерации как папку, ищем в имени файла
    if iterations is None:
        file_name = path_parts[-1]
        match = re.search(r'[_\s]?([0-9]+)\.[Hh][Tt][Mm]', file_name)
        if match:
            iterations = int(match.group(1))

    return {
        'projections': projections_folder,
        'method': method,
        'iterations': iterations
    }


def find_row_in_sheet(sheet, projections_folder, method, iterations):
    """
    Находит строку в листе Excel по проекциям, методу и итерациям
    """
    for row in range(4, sheet.max_row + 1):
        cell_projections = sheet.cell(row=row, column=4).value  # Колонка D - Угол
        cell_method = sheet.cell(row=row, column=5).value  # Колонка E - Метод
        cell_iterations = sheet.cell(row=row, column=6).value  # Колонка F - Количество итераций

        # Приводим к строке для сравнения
        if (str(cell_projections).strip() == str(projections_folder).strip() and
                str(cell_method).strip().upper() == str(method).strip().upper() and
                int(cell_iterations) == int(iterations)):
            return row

    return None


def process_and_fill_excel(base_directory, template_path, output_path):
    """
    Обрабатывает HTML файлы и заполняет Excel шаблон
    """
    print("=" * 80)
    print("Извлечение IQFinv и Threshold thickness из HTML и заполнение Excel-шаблона")
    print("=" * 80)
    print(f"Папка с данными: {base_directory}")
    print(f"Шаблон: {template_path}")
    print("-" * 80)

    # Проверяем существование шаблона
    if not os.path.exists(template_path):
        print(f"❌ ОШИБКА: Шаблон '{template_path}' не найден!")
        return

    # Загружаем шаблон Excel
    print("\n📂 Загрузка шаблона...")
    wb = load_workbook(template_path)
    sheet = wb.active

    # Ищем все .htm файлы
    html_files = glob.glob(os.path.join(base_directory, '**', '*.htm'), recursive=True)
    html_files.extend(glob.glob(os.path.join(base_directory, '**', '*.html'), recursive=True))

    print(f"🔍 Найдено HTML файлов: {len(html_files)}")
    print("-" * 80)

    processed_count = 0
    not_found_count = 0
    no_iqfinv_count = 0
    no_threshold_count = 0
    no_metadata_count = 0

    for html_file in sorted(html_files):
        # Извлекаем метаданные из пути
        metadata = extract_metadata(html_file)

        if not metadata['projections'] or not metadata['method'] or metadata['iterations'] is None:
            no_metadata_count += 1
            print(f"✗ Не удалось извлечь метаданные: {html_file}")
            print(
                f"  projections: {metadata['projections']}, method: {metadata['method']}, iterations: {metadata['iterations']}")
            continue

        # Находим строку в Excel
        row_num = find_row_in_sheet(sheet, metadata['projections'], metadata['method'], metadata['iterations'])

        if not row_num:
            not_found_count += 1
            print(
                f"✗ Не найдена строка в шаблоне: {metadata['projections']} / {metadata['method']} / {metadata['iterations']}")
            continue

        # Извлекаем IQFinv
        iqfinv = extract_iqfinv(html_file)
        if iqfinv is not None:
            sheet.cell(row=row_num, column=7, value=round(iqfinv, 1))  # Колонка G
            sheet.cell(row=row_num, column=7).alignment = Alignment(horizontal="center", vertical="center")
            print(f"  ✓ IQFinv: {round(iqfinv, 1)}", end="")
        else:
            no_iqfinv_count += 1
            print(f"  ✗ IQFinv не найден", end="")

        # Извлекаем значения толщины
        threshold_values = extract_threshold_values(html_file)
        if threshold_values:
            # Записываем значения толщины в ячейки (начиная с колонки H)
            for i, value in enumerate(threshold_values):
                sheet.cell(row=row_num, column=8 + i, value=float(value))
                sheet.cell(row=row_num, column=8 + i).alignment = Alignment(horizontal="center", vertical="center")
            print(f" | Threshold: {len(threshold_values)} значений", end="")
            processed_count += 1
        else:
            no_threshold_count += 1
            print(f" | ✗ Threshold не найден", end="")

        print(f" → {metadata['projections']} / {metadata['method']} / {metadata['iterations']}")

    # Сохраняем результат
    wb.save(output_path)

    print("\n" + "=" * 80)
    print(f"✅ Результат сохранён: {output_path}")
    print(f"✅ Обработано файлов: {processed_count}")
    print(f"⚠️ Не найдено в шаблоне: {not_found_count}")
    print(f"⚠️ Без IQFinv: {no_iqfinv_count}")
    print(f"⚠️ Без Threshold: {no_threshold_count}")
    print(f"⚠️ Без метаданных: {no_metadata_count}")
    print("=" * 80)


if __name__ == "__main__":
    base_dir = r'D:\Маммо_DBT\DBT_processing\no_grid_denoise_vraster' #Тут менять папку для обработки
    template_path = r'D:\Маммо_DBT\шаблон.xlsx'
    output_path = base_dir + r'\Шаблон_с_IQFinv_и_таблицами.xlsx'

    if not os.path.exists(base_dir):
        print(f"Ошибка: Директория не найдена: {base_dir}")
    else:
        process_and_fill_excel(base_dir, template_path, output_path)