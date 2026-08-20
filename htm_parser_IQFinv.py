import os
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment


def extract_iqfinv_simple(file_path):
    """Извлекает IQFinv из HTML файла"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        # Основной паттерн
        pattern = r'<td align="left">\s*IQFinv:\s*</td>\s*<td align="left">\s*([\d.]+)\s*</td>'
        match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
        if match:
            return float(match.group(1).strip())
        # Альтернативный
        alt_pattern = r'IQFinv\s*:\s*([\d.]+)'
        alt_match = re.search(alt_pattern, content, re.IGNORECASE)
        return float(alt_match.group(1).strip()) if alt_match else None
    except:
        return None


def parse_path_to_params(file_path):
    """Извлекает параметры из пути к файлу"""
    path_parts = Path(file_path).parts

    angle = None
    method = None
    iterations = None

    # Ищем папку с углом
    for i, part in enumerate(path_parts):
        if '_grad_' in part and '_projections' in part:
            angle = part
            # Метод - следующая папка
            if i + 1 < len(path_parts):
                method = path_parts[i + 1]
            # Итерации - следующая папка после метода
            if i + 2 < len(path_parts):
                iter_str = path_parts[i + 2]
                if iter_str.isdigit():
                    iterations = int(iter_str)
            break

    return angle, method, iterations


def main():
    # Пути
    base_path = r'D:\Маммо_DBT\DBT_processing\no_grid_denoise_vraster'
    template_path = r'D:\Маммо_DBT\шаблон.xlsx'  # Ваш шаблон Excel (должен быть в той же папке, где запускаете скрипт)
    output_path = base_path + r'\Шаблон_с_IQFinv.xlsx'  # Выходной файл

    print("=" * 80)
    print("Извлечение IQFinv из HTML и заполнение Excel-шаблона")
    print("=" * 80)
    print(f"Папка с данными: {base_path}")
    print(f"Шаблон: {template_path}")
    print("-" * 80)

    # Проверяем существование шаблона
    if not Path(template_path).exists():
        print(f"❌ ОШИБКА: Шаблон '{template_path}' не найден!")
        print("Убедитесь, что файл шаблон.xlsx находится в текущей папке.")
        return

    # Загружаем шаблон Excel
    print("\n📂 Загрузка шаблона...")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Создаём карту ячеек из шаблона
    cell_map = {}
    print("📋 Чтение структуры шаблона...")

    for row in range(6, ws.max_row + 1):  # Начинаем с 6 строки (после заголовков)
        angle = ws.cell(row=row, column=4).value  # Колонка D
        method = ws.cell(row=row, column=5).value  # Колонка E
        iterations = ws.cell(row=row, column=6).value  # Колонка F

        if angle and method and iterations:
            try:
                iter_val = int(iterations) if isinstance(iterations, (int, float)) else int(iterations)
                key = (str(angle).strip(), str(method).strip(), iter_val)
                cell_map[key] = row
            except (ValueError, TypeError):
                continue

    print(f"✅ Найдено строк в шаблоне: {len(cell_map)}")

    # Сканируем HTML файлы и собираем IQFinv
    print("\n🔍 Поиск HTML файлов...")
    results = {}
    file_count = 0

    for html_file in Path(base_path).rglob('*.htm'):
        iqfinv = extract_iqfinv_simple(html_file)
        if iqfinv is not None:
            # Извлекаем параметры из пути
            angle, method, iterations = parse_path_to_params(str(html_file))

            if angle and method and iterations:
                key = (angle, method, iterations)
                if key not in results:
                    results[key] = []
                results[key].append(iqfinv)
                file_count += 1
                print(f"  Найдено: {iqfinv} -> {angle} | {method} | {iterations}")

    print(f"\n✅ Всего найдено HTML файлов с IQFinv: {file_count}")
    print(f"✅ Уникальных комбинаций: {len(results)}")

    # Заполняем Excel
    print("\n✏️ Заполнение Excel...")
    filled_count = 0
    not_found_count = 0

    for (angle, method, iterations), values in results.items():
        # Берём среднее, если несколько значений
        avg_iqfinv = sum(values) / len(values)
        rounded_iqfinv = round(avg_iqfinv, 1)

        if (angle, method, iterations) in cell_map:
            row = cell_map[(angle, method, iterations)]
            cell = ws.cell(row=row, column=7, value=rounded_iqfinv)  # Колонка G
            cell.alignment = Alignment(horizontal="center", vertical="center")
            filled_count += 1
            print(f"  ✅ Записано: {angle} | {method} | {iterations} → {rounded_iqfinv}")
        else:
            not_found_count += 1
            print(f"  ⚠️ Не найдено в шаблоне: {angle} | {method} | {iterations}")

    # Сохраняем результат
    wb.save(output_path)

    print("\n" + "=" * 80)
    print(f"✅ Результат сохранён: {output_path}")
    print(f"✅ Заполнено ячеек: {filled_count}")
    if not_found_count > 0:
        print(f"⚠️ Не найдено в шаблоне: {not_found_count}")
    print("=" * 80)


if __name__ == "__main__":
    main()